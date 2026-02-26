#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
pdf_a_excel.py
==============
Script unificado que exporta PDF a Excel.
Detecta automáticamente si el PDF contiene tablas o solo texto,
o permite elegir el modo manualmente.

Requisitos:
    pip install tabula-py pdfplumber openpyxl pandas

Uso:
    python pdf_a_excel.py entrada.pdf
    python pdf_a_excel.py entrada.pdf -o salida.xlsx --modo tablas
    python pdf_a_excel.py entrada.pdf -o salida.xlsx --modo texto --separador ";"
    python pdf_a_excel.py entrada.pdf --modo auto --paginas 1-10
    python pdf_a_excel.py carpeta_pdfs/ -o carpeta_salida/ --batch
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter

# ---------- Verificación de dependencias ----------
_dependencias_faltantes = []

try:
    import pdfplumber
except ImportError:
    pdfplumber = None
    _dependencias_faltantes.append("pdfplumber")

try:
    import tabula
except ImportError:
    tabula = None
    _dependencias_faltantes.append("tabula-py")

try:
    import openpyxl  # noqa: F401
except ImportError:
    _dependencias_faltantes.append("openpyxl")

if _dependencias_faltantes:
    print("Dependencias faltantes:")
    for dep in _dependencias_faltantes:
        print(f"  - {dep}")
    print(f"\nInstálalas con:  pip install {' '.join(_dependencias_faltantes)}")
    if tabula is None:
        print("Nota: tabula-py requiere Java instalado.")
    sys.exit(1)


# =====================================================================
# Argumentos
# =====================================================================
def parsear_argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Exporta PDF a Excel (.xlsx). Soporta tablas y texto.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Ejemplos:\n"
            "  python pdf_a_excel.py factura.pdf\n"
            "  python pdf_a_excel.py reporte.pdf -o reporte.xlsx --modo tablas --lattice\n"
            "  python pdf_a_excel.py libro.pdf --modo texto --sin-vacias\n"
            "  python pdf_a_excel.py pdfs/ --batch -o excels/\n"
        ),
    )

    # --- Entrada / Salida ---
    parser.add_argument(
        "entrada",
        type=str,
        help="Ruta al PDF de entrada, o carpeta con PDFs si se usa --batch.",
    )
    parser.add_argument(
        "-o", "--output",
        type=str,
        default=None,
        help="Ruta del Excel de salida (o carpeta de salida con --batch).",
    )

    # --- Modo ---
    parser.add_argument(
        "-m", "--modo",
        choices=["auto", "tablas", "texto"],
        default="auto",
        help=(
            "Modo de extracción: 'auto' detecta tablas primero, "
            "'tablas' fuerza extracción de tablas, "
            "'texto' fuerza extracción de texto. Por defecto: auto"
        ),
    )

    # --- Páginas ---
    parser.add_argument(
        "-p", "--paginas",
        type=str,
        default="all",
        help="Páginas a procesar: 'all', '1', '1,3,5', '1-5'. Por defecto: all",
    )

    # --- Opciones de tablas ---
    grupo_tablas = parser.add_argument_group("Opciones de tablas")
    grupo_tablas.add_argument(
        "--lattice",
        action="store_true",
        default=False,
        help="Modo lattice: tablas con bordes visibles.",
    )
    grupo_tablas.add_argument(
        "--stream",
        action="store_true",
        default=False,
        help="Modo stream: tablas sin bordes.",
    )
    grupo_tablas.add_argument(
        "--multiple-tablas",
        action="store_true",
        default=False,
        help="Detectar múltiples tablas por página.",
    )
    grupo_tablas.add_argument(
        "--separar-hojas",
        action="store_true",
        default=False,
        help="Guardar cada tabla en una hoja separada (por defecto todo va en una sola hoja).",
    )

    # --- Opciones de texto ---
    grupo_texto = parser.add_argument_group("Opciones de texto")
    grupo_texto.add_argument(
        "--separador",
        type=str,
        default=None,
        help="Separador para dividir líneas en columnas (ej: ';', ',', '|').",
    )
    grupo_texto.add_argument(
        "--sin-vacias",
        action="store_true",
        default=False,
        help="Omitir líneas vacías al extraer texto.",
    )
    grupo_texto.add_argument(
        "--modo-texto",
        choices=["linea", "pagina"],
        default="linea",
        help="'linea' o 'pagina'. Por defecto: linea",
    )

    # --- Batch ---
    parser.add_argument(
        "--batch",
        action="store_true",
        default=False,
        help="Procesar todos los PDFs de una carpeta.",
    )

    # --- Otros ---
    parser.add_argument(
        "--encoding",
        type=str,
        default="utf-8",
        help="Codificación. Por defecto: utf-8",
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        default=False,
        help="Mostrar información detallada.",
    )

    return parser.parse_args()


# =====================================================================
# Utilidades de páginas
# =====================================================================
def parse_rango_paginas(paginas_str: str, total_paginas: int) -> list[int]:
    if paginas_str.strip().lower() == "all":
        return list(range(total_paginas))
    indices = set()
    for parte in paginas_str.split(","):
        parte = parte.strip()
        if "-" in parte:
            ini, fin = parte.split("-", 1)
            for p in range(int(ini), int(fin) + 1):
                if 1 <= p <= total_paginas:
                    indices.add(p - 1)
        else:
            p = int(parte)
            if 1 <= p <= total_paginas:
                indices.add(p - 1)
    return sorted(indices)


# =====================================================================
# Extracción de tablas (tabula-py)
# =====================================================================
def extraer_tablas_tabula(pdf_path: str, args: argparse.Namespace) -> list[pd.DataFrame]:
    """Extrae tablas usando tabula-py (requiere Java)."""
    kwargs = {
        "input_path": pdf_path,
        "pages": args.paginas,
        "multiple_tables": True,  # Siempre True para evitar errores de columnas irregulares
        "encoding": args.encoding,
    }
    if args.lattice:
        kwargs["lattice"] = True
    if args.stream:
        kwargs["stream"] = True

    if args.verbose:
        print(f"  [tabula] Extrayendo tablas (páginas={args.paginas}) ...")

    try:
        tablas = tabula.read_pdf(**kwargs)
        resultado = [t for t in tablas if not t.empty]
        if resultado and args.verbose:
            print(f"  [tabula] {len(resultado)} tabla(s) extraídas.")
        return resultado
    except Exception as e:
        if args.verbose:
            print(f"  [tabula] Error: {e}")
        return []


# =====================================================================
# Extracción de tablas (pdfplumber) — alternativa sin Java
# =====================================================================
def extraer_tablas_pdfplumber(pdf_path: str, args: argparse.Namespace) -> list[pd.DataFrame]:
    """Extrae tablas usando pdfplumber (no requiere Java)."""
    if args.verbose:
        print(f"  [pdfplumber] Extrayendo tablas ...")

    tablas = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            total = len(pdf.pages)
            paginas_idx = parse_rango_paginas(args.paginas, total)

            for idx in paginas_idx:
                page_tables = pdf.pages[idx].extract_tables()
                if page_tables:
                    for raw_table in page_tables:
                        if raw_table and len(raw_table) > 0:
                            # Primera fila como cabecera, resto como datos
                            header = raw_table[0]
                            data = raw_table[1:]
                            if data:
                                df = pd.DataFrame(data, columns=header)
                                # Limpiar columnas y filas vacías
                                df = df.dropna(how="all", axis=0).dropna(how="all", axis=1)
                                if not df.empty:
                                    tablas.append(df)

        if tablas and args.verbose:
            print(f"  [pdfplumber] {len(tablas)} tabla(s) extraídas.")
    except Exception as e:
        if args.verbose:
            print(f"  [pdfplumber] Error extrayendo tablas: {e}")

    return tablas


# =====================================================================
# Extracción de texto (pdfplumber)
# =====================================================================
def extraer_texto(pdf_path: str, args: argparse.Namespace) -> pd.DataFrame:
    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        paginas_idx = parse_rango_paginas(args.paginas, total)

        if args.verbose:
            print(f"  [pdfplumber-texto] {total} páginas, procesando {len(paginas_idx)}.")

        filas = []
        for idx in paginas_idx:
            texto = pdf.pages[idx].extract_text() or ""
            if args.modo_texto == "pagina":
                filas.append({"Pagina": idx + 1, "Texto": texto})
            else:
                for n, linea in enumerate(texto.split("\n"), 1):
                    if args.sin_vacias and not linea.strip():
                        continue
                    filas.append({"Pagina": idx + 1, "Linea": n, "Texto": linea})

    if not filas:
        return pd.DataFrame()

    df = pd.DataFrame(filas)

    if args.separador and "Texto" in df.columns:
        sep = args.separador.replace("\\t", "\t")
        extras = df["Texto"].str.split(sep, expand=True)
        extras.columns = [f"Col_{i+1}" for i in range(extras.shape[1])]
        df = pd.concat([df.drop(columns=["Texto"]), extras], axis=1)

    return df


# =====================================================================
# Auto-detección y conversión de tipos
# =====================================================================
_PATRON_FECHA = re.compile(
    r'^\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}$'
)
_FORMATOS_FECHA = [
    "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y",
    "%d/%m/%y", "%m/%d/%y", "%d-%m-%y", "%m-%d-%y",
    "%Y-%m-%d", "%Y/%m/%d",
]


def _parsear_fecha(val: str) -> datetime | None:
    """Intenta parsear una cadena como fecha."""
    val = val.strip()
    if not _PATRON_FECHA.match(val):
        return None
    for fmt in _FORMATOS_FECHA:
        try:
            return datetime.strptime(val, fmt)
        except ValueError:
            continue
    return None


def _parsear_numero(val: str) -> float | None:
    """Intenta parsear una cadena como número (soporta comas como miles)."""
    val = val.strip()
    if not val:
        return None
    # Quitar espacios internos
    val = val.replace(" ", "")
    # Detectar formato: 1,234.56 o 1.234,56 o simples
    # Formato americano: comas como miles, punto decimal
    if re.match(r'^-?[\d,]+\.\d+$', val):
        try:
            return float(val.replace(",", ""))
        except ValueError:
            return None
    # Formato europeo: puntos como miles, coma decimal
    if re.match(r'^-?[\d.]+,\d+$', val):
        try:
            return float(val.replace(".", "").replace(",", "."))
        except ValueError:
            return None
    # Número entero con comas como miles: 140,000
    if re.match(r'^-?[\d,]+$', val) and "," in val:
        try:
            return float(val.replace(",", ""))
        except ValueError:
            return None
    # Número simple
    if re.match(r'^-?\d+\.?\d*$', val):
        try:
            return float(val)
        except ValueError:
            return None
    return None


def auto_convertir_tipos(df: pd.DataFrame) -> pd.DataFrame:
    """Convierte columnas de texto a número o fecha cuando la mayoría lo son."""
    df = df.copy()
    for col in df.columns:
        if df[col].dtype != object:
            continue

        # Tomar muestra de valores no vacíos
        muestra = df[col].dropna()
        muestra = muestra[muestra.astype(str).str.strip() != ""]
        if muestra.empty:
            continue

        # --- Intentar fechas (solo si mayoría son fechas) ---
        fechas = muestra.astype(str).apply(_parsear_fecha)
        ratio_fechas = fechas.notna().sum() / len(muestra) if len(muestra) > 0 else 0
        if ratio_fechas >= 0.6 and fechas.notna().sum() >= 1:
            df[col] = df[col].astype(str).apply(
                lambda x: _parsear_fecha(x) if pd.notna(x) and str(x).strip() else None
            )
            df[col] = pd.to_datetime(df[col], errors="coerce")
            continue

        # --- Intentar números (solo si mayoría son números) ---
        nums = muestra.astype(str).apply(_parsear_numero)
        ratio_nums = nums.notna().sum() / len(muestra) if len(muestra) > 0 else 0
        if ratio_nums >= 0.6 and nums.notna().sum() >= 1:
            df[col] = df[col].astype(str).apply(
                lambda x: _parsear_numero(x) if pd.notna(x) and str(x).strip() else None
            )
            continue

    return df


# =====================================================================
# Aplicar formatos al Excel
# =====================================================================
def _aplicar_formatos_hoja(ws):
    """Aplica formato numérico, fecha y ancho de columna a una hoja.
    Convierte celda por celda los strings que son números o fechas."""
    # Formato encabezado
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Detectar tipos por columna y aplicar formato
    for col_idx in range(1, ws.max_column + 1):
        max_ancho = len(str(ws.cell(1, col_idx).value or ""))
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            val = cell.value
            if val is None:
                continue

            # --- Si es string, intentar convertir celda por celda ---
            if isinstance(val, str):
                s = val.strip()
                if s:
                    # Intentar fecha
                    fecha = _parsear_fecha(s)
                    if fecha is not None:
                        cell.value = fecha
                        val = fecha
                    else:
                        # Intentar número
                        num = _parsear_numero(s)
                        if num is not None:
                            cell.value = num
                            val = num

            # --- Aplicar formato según tipo final ---
            if isinstance(val, datetime):
                cell.number_format = "DD/MM/YYYY"
                cell.alignment = Alignment(horizontal="center")
                max_ancho = max(max_ancho, 12)
            elif isinstance(val, (int, float)):
                if isinstance(val, float) and val != int(val):
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
                max_ancho = max(max_ancho, len(f"{val:,.2f}"))
            else:
                max_ancho = max(max_ancho, min(len(str(val)), 50))

        # Auto-ancho (con un máximo de 55)
        ancho = min(max_ancho + 3, 55)
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho


# =====================================================================
# Guardar en Excel
# =====================================================================
def guardar_tablas_excel(tablas: list[pd.DataFrame], ruta: str, separar_hojas: bool):
    # Convertir tipos en cada tabla
    tablas = [auto_convertir_tipos(t) for t in tablas]

    with pd.ExcelWriter(ruta, engine="openpyxl") as w:
        if separar_hojas:
            for i, df in enumerate(tablas, 1):
                nombre = f"Tabla_{i}"[:31]
                df.to_excel(w, sheet_name=nombre, index=False)
        else:
            pd.concat(tablas, ignore_index=True).to_excel(w, sheet_name="Datos", index=False)

        # Aplicar formatos a cada hoja
        for ws in w.book.worksheets:
            _aplicar_formatos_hoja(ws)


def guardar_texto_excel(df: pd.DataFrame, ruta: str):
    df = auto_convertir_tipos(df)
    with pd.ExcelWriter(ruta, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Texto")
        _aplicar_formatos_hoja(w.book["Texto"])


# =====================================================================
# Procesamiento de un solo PDF
# =====================================================================
def procesar_pdf(pdf_path: str, ruta_salida: str, args: argparse.Namespace):
    print(f"\nProcesando: {pdf_path}")

    modo = args.modo
    tablas = []

    # --- Extracción de tablas ---
    if modo in ("auto", "tablas"):
        # 1) Intentar con tabula-py
        tablas = extraer_tablas_tabula(pdf_path, args)

        # 2) Fallback: pdfplumber extract_tables
        if not tablas:
            if args.verbose:
                print("  [fallback] tabula no encontró tablas, probando pdfplumber ...")
            tablas = extraer_tablas_pdfplumber(pdf_path, args)

        if tablas:
            guardar_tablas_excel(tablas, ruta_salida, args.separar_hojas)
            total_filas = sum(len(t) for t in tablas)
            print(f"  OK: {len(tablas)} tabla(s), {total_filas} filas -> {ruta_salida}")
            return

        # Si forzamos tablas y no hay, avisar
        if modo == "tablas":
            print("  Sin tablas encontradas.")
            return

        # modo auto: caer a texto
        print("  No se detectaron tablas, extrayendo texto ...")

    # --- Extracción de texto ---
    df = extraer_texto(pdf_path, args)
    if df.empty:
        print("  Sin texto extraído.")
        return
    guardar_texto_excel(df, ruta_salida)
    print(f"  OK: {len(df)} registros de texto -> {ruta_salida}")


# =====================================================================
# Main
# =====================================================================
def main():
    args = parsear_argumentos()
    entrada = Path(args.entrada)

    if args.batch:
        # --- Modo batch: procesar carpeta ---
        if not entrada.is_dir():
            sys.exit(f"Error: '{entrada}' no es una carpeta (requerido con --batch).")

        pdfs = sorted(entrada.glob("*.pdf"))
        if not pdfs:
            sys.exit(f"No se encontraron archivos .pdf en '{entrada}'.")

        carpeta_salida = Path(args.output) if args.output else entrada / "excel_output"
        carpeta_salida.mkdir(parents=True, exist_ok=True)

        print(f"Modo batch: {len(pdfs)} PDFs encontrados.")
        errores = []
        for pdf_file in pdfs:
            salida = carpeta_salida / (pdf_file.stem + ".xlsx")
            try:
                procesar_pdf(str(pdf_file), str(salida), args)
            except Exception as e:
                print(f"  ERROR en '{pdf_file.name}': {e}")
                errores.append(pdf_file.name)

        if errores:
            print(f"\n{len(errores)} archivo(s) con errores: {', '.join(errores)}")
        print(f"\nBatch completado. Archivos en: {carpeta_salida}")

    else:
        # --- Modo individual ---
        if not entrada.is_file():
            sys.exit(f"Error: No se encontró el archivo '{entrada}'.")

        if args.output:
            ruta_salida = args.output
        else:
            ruta_salida = str(entrada.with_suffix(".xlsx"))

        procesar_pdf(str(entrada), ruta_salida, args)

    print("\n¡Listo!")


if __name__ == "__main__":
    main()
